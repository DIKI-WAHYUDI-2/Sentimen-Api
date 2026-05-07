from pathlib import Path
import re
from collections import Counter
from openpyxl import load_workbook, Workbook

INPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita.xlsx")
OUTPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\review_label_negatif.xlsx")
SHEET_NAME = "DATASET"

POSITIVE_CUES = [
    'apresiasi', 'penghargaan', 'sertifikat', 'sertifikasi', 'dukung', 'komitmen', 'berhasil', 'sukses',
    'manfaat', 'bakti sosial', 'bantuan', 'sembako', 'stunting', 'pasar murah', 'salurkan', 'sinergi',
    'kolaborasi', 'ekspor', 'beasiswa', 'juara', 'prestasi', 'naik kelas', 'sertifikasi', 'kinerja gemilang',
    'motor penggerak', 'ketahanan pangan', 'energi hijau', 'digitalisasi', 'zero fatality', 'anti penyuapan',
    'anti korupsi', 'meringankan beban', 'peduli', 'mengapresiasi'
]

MARKET_NEUTRAL_CUES = [
    'harga cpo', 'harga sawit', 'harga tbs', 'harga referensi', 'bea keluar', 'pungutan ekspor', 'stok cpo',
    'produksi cpo', 'kontrak berjangka', 'bursa malaysia', 'ringgit', 'minyak mentah', 'permintaan', 'ekspor kuat',
    'prediksi', 'surveyor kargo', 'pdb q3', 'pasar bersikap hati hati', 'harga pembelian tbs', 'disbun riau'
]

LEGAL_CONFLICT_CUES = [
    'gugatan', 'digugat', 'menggugat', 'wanprestasi', 'sengketa', 'konflik', 'demo', 'tuntutan',
    'bermasalah', 'persidangan', 'sidang', 'pengadilan', 'selisih timbangan', 'ganti rugi', 'koppsa', 'apkasindo'
]

STRONG_NEGATIVE_CUES = [
    'korupsi', 'ilegal', 'melanggar', 'pelanggaran', 'merugikan', 'gagal', 'pencemaran', 'pengusiran paksa',
    'hancur kebun', 'diduga penyebab', 'dituding', 'disalahkan', 'krisis', 'kebakaran', 'suap', 'fraud'
]

PTPN_DEFENSIVE_CUES = [
    'kuasa hukum ptpn', 'ptpn menggugat', 'gugatan ptpn', 'ptpn menyambut', 'ptpn mengapresiasi',
    'ptpn salurkan', 'ptpn membantu', 'ptpn regional iii terhadap', 'menyambut kehadiran massa aksi secara terbuka',
    'dinyatakan lemah', 'diterima dengan baik', 'apresiasi kepada ptpn'
]



def norm(text: str) -> str:
    text = (text or '').lower()
    text = text.replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()



def contains_any(text: str, cues: list[str]) -> bool:
    return any(cue in text for cue in cues)



def review_negative(title: str, content: str) -> tuple[str, str, str]:
    t = norm(title)
    c = norm(content)
    combined = f'{t} {c}'.strip()

    if not combined:
        return 'Meragukan', 'netral', 'Tidak ada konteks yang cukup untuk memastikan PTPN benar-benar digambarkan negatif.'

    # Positive/supportive coverage that was mislabeled negative.
    if contains_any(combined, POSITIVE_CUES) and not contains_any(combined, STRONG_NEGATIVE_CUES):
        return 'Salah', 'positif', 'Isi berita justru menampilkan dukungan, bantuan, apresiasi, atau pencapaian PTPN.'

    # Market and commodity movement is not automatically sentiment toward PTPN.
    if contains_any(combined, MARKET_NEUTRAL_CUES):
        if contains_any(combined, ['ptpn v sei buatan', 'ptpn v sei tapung']) and ('turun' in combined or 'penurunan' in combined):
            return 'Meragukan', 'netral', 'Berita membahas penurunan harga pasar/TBS dan hanya sebagian menyebut unit PTPN, tanpa penilaian kuat terhadap PTPN.'
        return 'Salah', 'netral', 'Berita terutama membahas pergerakan harga atau pasar CPO/TBS, bukan penilaian negatif terhadap PTPN.'

    # Cases where PTPN is clearly criticized or tied to wrongdoing.
    if contains_any(combined, STRONG_NEGATIVE_CUES):
        if contains_any(combined, PTPN_DEFENSIVE_CUES):
            return 'Meragukan', 'netral', 'Ada unsur negatif, tetapi konteks berita menunjukkan PTPN tidak secara jelas menjadi pihak yang disalahkan.'
        if 'ptpn' in combined:
            return 'Jelas Negatif', 'negatif', 'Berita mengaitkan PTPN dengan tuduhan, pelanggaran, atau hal buruk secara cukup jelas.'

    # Legal/conflict coverage is often ambiguous from PTPN perspective.
    if contains_any(combined, LEGAL_CONFLICT_CUES):
        if contains_any(combined, PTPN_DEFENSIVE_CUES):
            return 'Salah', 'positif', 'Meski bertema sengketa, isi berita cenderung membela atau menguntungkan posisi PTPN.'
        return 'Meragukan', 'netral', 'Berita bertema sengketa atau proses hukum, tetapi posisi salah-benar PTPN tidak cukup tegas.'

    # Bencana/korban should not be auto-negative if PTPN is helping.
    if ('banjir' in combined or 'korban' in combined or 'bencana' in combined) and contains_any(combined, ['bantuan', 'salurkan', 'sembako', 'meringankan beban', 'peduli']):
        return 'Salah', 'positif', 'Kejadian buruk dibahas, tetapi PTPN justru hadir membantu sehingga sentimennya tidak negatif bagi PTPN.'

    # News where title is negative but article does not sustain the negative view.
    if 'turun' in combined or 'penurunan' in combined or 'tekanan' in combined or 'anjlok' in combined:
        return 'Meragukan', 'netral', 'Ada kosa kata negatif, tetapi konteksnya lebih informatif atau sektoral daripada menyudutkan PTPN.'

    return 'Meragukan', 'netral', 'Konteks negatif terhadap PTPN tidak cukup tegas, sehingga lebih aman dikategorikan meragukan.'



def main() -> None:
    wb = load_workbook(INPUT_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Review Negatif'
    out_ws.append([
        'ROW_ASAL', 'Judul', 'Link', 'ISI BERITA', 'LABEL_ASAL',
        'HASIL_REVIEW', 'LABEL_DISARANKAN', 'ALASAN_REVIEW'
    ])

    counts = Counter()
    total = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = row[0] or ''
        link = row[1] or ''
        content = row[2] or ''
        original_label = row[3]
        if original_label != 'negatif':
            continue

        total += 1
        review_cat, suggested_label, reason = review_negative(title, content)
        counts[review_cat] += 1
        out_ws.append([
            row_idx, title, link, content, original_label,
            review_cat, suggested_label, reason
        ])

    summary = out_wb.create_sheet('Ringkasan')
    summary.append(['METRIK', 'JUMLAH'])
    summary.append(['Total data negatif diperiksa', total])
    summary.append(['Jelas Negatif', counts['Jelas Negatif']])
    summary.append(['Meragukan', counts['Meragukan']])
    summary.append(['Salah', counts['Salah']])

    out_wb.save(OUTPUT_PATH)
    print(f'saved={OUTPUT_PATH}')
    print(f'total={total}')
    print(f'jelas_negatif={counts["Jelas Negatif"]}')
    print(f'meragukan={counts["Meragukan"]}')
    print(f'salah={counts["Salah"]}')


if __name__ == '__main__':
    main()
