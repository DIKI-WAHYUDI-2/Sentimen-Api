from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita.xlsx")
SHEET_NAME = "Data"

# Pedoman labelisasi:
# - positif: capaian, bantuan, kolaborasi, apresiasi, program, penghargaan, penguatan kinerja/citra PTPN
# - netral: berita informatif/deskriptif tanpa evaluasi kuat terhadap PTPN
# - negatif: berita yang menonjolkan masalah/risiko/penurunan/konotasi buruk terhadap PTPN atau isu sawit terkait
NEGATIVE_ROWS = {
    34,  # dampak negatif sawit / pengusaha ilegal
    66,  # antisipasi penurunan pendapatan
}

NEUTRAL_ROWS = {
    11,
    12,
    17,
    21,
    32,
    37,
    39,
    41,
    46,
    63,
    64,
    65,
    71,
    74,
    76,
    80,
    81,
    84,
    85,
    87,
    88,
    89,
    90,
    100,
    107,
    112,
    125,
    132,
}


def label_for_row(row_number: int) -> str:
    if row_number in NEGATIVE_ROWS:
        return "negatif"
    if row_number in NEUTRAL_ROWS:
        return "netral"
    return "positif"


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    if ws["A1"].value != "JUDUL PEMBERITAAN" or ws["C1"].value != "LABEL":
        raise ValueError("Struktur sheet 'Data' tidak sesuai ekspektasi.")

    counts = {"positif": 0, "netral": 0, "negatif": 0}

    for row_idx in range(2, ws.max_row + 1):
        if not any(ws.cell(row=row_idx, column=col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            continue
        label = label_for_row(row_idx)
        ws.cell(row=row_idx, column=3).value = label
        counts[label] += 1

    wb.save(WORKBOOK_PATH)

    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    for key in ("positif", "netral", "negatif"):
        print(f"{key}={counts[key]}")


if __name__ == "__main__":
    main()
