from pathlib import Path
import re
from collections import Counter

from openpyxl import load_workbook

INPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita.xlsx")
FALLBACK_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita-labeled.xlsx")
SHEET_NAME = "Data"

POSITIVE_PATTERNS = [
    r"\bapresiasi\b", r"\braih\b", r"\bmeraih\b", r"\bpenghargaan\b", r"\bsertifikat\b",
    r"\bdukung\w*\b", r"\bkomitmen\b", r"\bberhasil\b", r"\bsukses\b", r"\boptimal\b",
    r"\btingkatkan\w*\b", r"\bmeningkat\w*\b", r"\bnaik\b", r"\bmelonjak\b", r"\bmelesat\b",
    r"\bmelambung\b", r"\bekspor\b", r"\bbeasiswa\b", r"\bbakti sosial\b", r"\bsalurkan\w*\b",
    r"\bbantuan\b", r"\bstunting\b", r"\bsehat\w*\b", r"\bcermat\b", r"\bgrand launching\b",
    r"\bperkuat\w*\b", r"\bperesmian\b", r"\bluncurkan\w*\b", r"\bkolaborasi\b", r"\bsinergi\b",
    r"\bprestasi\b", r"\bjuara\b", r"\bpositif\b", r"\bmanfaat\b", r"\bketahanan pangan\b",
    r"\bswasembada\b", r"\brevitalisasi\b", r"\btransformasi\b", r"\bdigitalisasi\b", r"\bparipurna\b",
    r"\butama\b", r"\bberkembang\b", r"\bvaluasi capai\b", r"\btargetkan\b", r"\bgercep\b",
]

NEGATIVE_PATTERNS = [
    r"\banjlok\w*\b", r"\bpenurunan\b", r"\bmenurun\b", r"\bturun\b", r"\btekanan\b",
    r"\bilegal\b", r"\bdampak negatif\b", r"\bnegatif\b", r"\bkorupsi\b", r"\bkrisis\b",
    r"\bkebakaran\b", r"\bbanjir\b", r"\bkorban\b", r"\bgagal\b", r"\bsengketa\b",
    r"\bmasalah\b", r"\bkonflik\b", r"\bterlarang\b", r"\bpencemaran\b", r"\bmerosot\b",
]

NEUTRAL_PATTERNS = [
    r"\bhak jawab\b", r"\bklarifikasi\b", r"\bkunjungan\b", r"\baudiensi\b", r"\btarget\b",
    r"\binisiatif\b", r"\bprediksi\b", r"\bharga cpo\b", r"\bharga tbs\b", r"\bdaftar\b",
    r"\bbahas\b", r"\bpaparan\b", r"\bworkshop\b", r"\brapat\b", r"\bevaluasi\b",
]

POSITIVE_EXCEPTIONS = [
    r"anti penyuapan", r"anti korupsi", r"korban dampak banjir", r"bakti sosial", r"pasar murah",
    r"sembako", r"donor darah", r"makan bergizi", r"bantuan", r"peduli", r"cegah stunting",
]

NEGATIVE_TITLE_EXACT = {
    "Hak Jawab Klarifikasi dan Hak Jawab Pemberitaan": "netral",
}


def normalize(text: str) -> str:
    text = (text or "").lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def count_patterns(text: str, patterns: list[str]) -> int:
    total = 0
    for pat in patterns:
        if re.search(pat, text):
            total += 1
    return total


def classify(title: str, content: str) -> str:
    raw_title = (title or "").strip()
    if raw_title in NEGATIVE_TITLE_EXACT:
        return NEGATIVE_TITLE_EXACT[raw_title]

    title_n = normalize(title)
    content_n = normalize(content)
    combined = f"{title_n} {content_n}".strip()

    if not combined:
        return "netral"

    # Maintenance / extraction failure pages should rely more on title.
    if "kami sedang mengerjakan beberapa pekerjaan" in combined:
        combined = title_n

    pos = count_patterns(title_n, POSITIVE_PATTERNS) * 3 + count_patterns(content_n, POSITIVE_PATTERNS)
    neg = count_patterns(title_n, NEGATIVE_PATTERNS) * 3 + count_patterns(content_n, NEGATIVE_PATTERNS)
    neu = count_patterns(title_n, NEUTRAL_PATTERNS) * 2 + count_patterns(content_n, NEUTRAL_PATTERNS)

    if any(exc in combined for exc in POSITIVE_EXCEPTIONS):
        pos += 4
        neg = max(0, neg - 2)

    # Market/forecast news is usually better treated as neutral unless the polarity is very strong.
    if re.search(r"\bharga (cpo|tbs)\b", combined) or re.search(r"\bprediksi\b", combined):
        neu += 3

    # If title is dominated by direct praise/capability claims, keep it positive.
    if re.search(r"\bapresiasi\b|\braih\b|\bjuara\b|\bprestasi\b|\bpenghargaan\b|\bdukung\w*\b", title_n):
        pos += 2

    # Strongly negative sector/legal headlines.
    if re.search(r"\bilegal\b|\bdampak negatif\b|\bkorupsi\b|\banjlok\w*\b|\bpenurunan\b", title_n):
        neg += 3

    # Choose label conservatively.
    if neg >= pos + 2 and neg >= neu:
        return "negatif"
    if pos >= neg + 2 and pos >= neu:
        return "positif"
    return "netral"


def main() -> None:
    wb = load_workbook(INPUT_PATH)
    ws = wb[SHEET_NAME]
    counts = Counter()

    for row_idx in range(2, ws.max_row + 1):
        title = ws.cell(row=row_idx, column=1).value or ""
        content = ws.cell(row=row_idx, column=3).value or ""
        label = classify(title, content)
        ws.cell(row=row_idx, column=4).value = label
        counts[label] += 1

    try:
        wb.save(INPUT_PATH)
        saved = INPUT_PATH
    except PermissionError:
        wb.save(FALLBACK_PATH)
        saved = FALLBACK_PATH

    print(f"saved={saved}")
    for key in ("positif", "netral", "negatif"):
        print(f"{key}={counts[key]}")


if __name__ == "__main__":
    main()
