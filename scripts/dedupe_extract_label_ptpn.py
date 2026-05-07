from pathlib import Path
from collections import Counter
import re
import requests
import trafilatura
from openpyxl import Workbook, load_workbook

INPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita.xlsx")
OUTPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita-clean-labeled.xlsx")
SHEET_NAME = "Data"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/132.0.0.0 Safari/537.36"
)

POSITIVE_PATTERNS = [
    r"\bapresiasi\b", r"\braih\b", r"\bmeraih\b", r"\bpenghargaan\b", r"\bsertifikat\b",
    r"\bdukung\w*\b", r"\bkomitmen\b", r"\bberhasil\b", r"\bsukses\b", r"\boptimal\b",
    r"\btingkatkan\w*\b", r"\bmeningkat\w*\b", r"\bnaik\b", r"\bmelonjak\b", r"\bmelesat\b",
    r"\bmelambung\b", r"\bekspor\b", r"\bbeasiswa\b", r"\bbakti sosial\b", r"\bsalurkan\w*\b",
    r"\bbantuan\b", r"\bstunting\b", r"\bsehat\w*\b", r"\bgrand launching\b", r"\bperkuat\w*\b",
    r"\bperesmian\b", r"\bluncurkan\w*\b", r"\bkolaborasi\b", r"\bsinergi\b", r"\bprestasi\b",
    r"\bjuara\b", r"\bmanfaat\b", r"\bketahanan pangan\b", r"\bswasembada\b", r"\brevitalisasi\b",
    r"\btransformasi\b", r"\bdigitalisasi\b", r"\bparipurna\b", r"\bberkembang\b", r"\btargetkan\b",
]
NEGATIVE_PATTERNS = [
    r"\banjlok\w*\b", r"\bpenurunan\b", r"\bmenurun\b", r"\bturun\b", r"\btekanan\b",
    r"\bilegal\b", r"\bdampak negatif\b", r"\bnegatif\b", r"\bkrisis\b", r"\bkebakaran\b",
    r"\bbanjir\b", r"\bkorban\b", r"\bgagal\b", r"\bsengketa\b", r"\bmasalah\b",
    r"\bkonflik\b", r"\bpencemaran\b", r"\bmerosot\b",
]
NEUTRAL_PATTERNS = [
    r"\bhak jawab\b", r"\bklarifikasi\b", r"\bkunjungan\b", r"\baudiensi\b", r"\btarget\b",
    r"\binisiatif\b", r"\bprediksi\b", r"\bharga cpo\b", r"\bharga tbs\b", r"\bdaftar\b",
    r"\bbahas\b", r"\bpaparan\b", r"\bworkshop\b", r"\brapat\b", r"\bevaluasi\b",
]
POSITIVE_EXCEPTIONS = [
    "anti penyuapan", "anti korupsi", "korban dampak banjir", "bakti sosial", "pasar murah",
    "sembako", "donor darah", "makan bergizi", "bantuan", "peduli", "cegah stunting",
]


def normalize_whitespace(text: str) -> str:
    text = text or ""
    text = text.replace("\r", "\n")
    lines = [line.strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    return "\n\n".join(lines).strip()


def normalize_key(text: str) -> str:
    text = (text or "").lower()
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def content_key(text: str) -> str:
    text = normalize_key(text)
    if not text:
        return ""
    words = text.split()
    return " ".join(words[:120])


def count_patterns(text: str, patterns: list[str]) -> int:
    return sum(1 for p in patterns if re.search(p, text))


def classify(title: str, content: str) -> str:
    title_n = normalize_key(title)
    content_n = normalize_key(content)
    combined = (title_n + " " + content_n).strip()
    if not combined:
        return "netral"
    pos = count_patterns(title_n, POSITIVE_PATTERNS) * 3 + count_patterns(content_n, POSITIVE_PATTERNS)
    neg = count_patterns(title_n, NEGATIVE_PATTERNS) * 3 + count_patterns(content_n, NEGATIVE_PATTERNS)
    neu = count_patterns(title_n, NEUTRAL_PATTERNS) * 2 + count_patterns(content_n, NEUTRAL_PATTERNS)
    if any(x in combined for x in POSITIVE_EXCEPTIONS):
        pos += 4
        neg = max(0, neg - 2)
    if re.search(r"\bharga (cpo|tbs|sawit)\b", combined) or re.search(r"\bprediksi\b", combined):
        neu += 3
    if re.search(r"\banjlok\w*\b|\bpenurunan\b|\bdampak negatif\b|\bilegal\b", title_n):
        neg += 3
    if neg >= pos + 2 and neg >= neu:
        return "negatif"
    if pos >= neg + 2 and pos >= neu:
        return "positif"
    return "netral"


def fetch_article(url: str) -> tuple[str, str]:
    try:
        response = requests.get(url, timeout=30, headers={"User-Agent": USER_AGENT}, allow_redirects=True)
        response.raise_for_status()
    except Exception as exc:
        return "", f"request_error:{type(exc).__name__}"

    extracted = trafilatura.extract(
        response.text,
        url=url,
        include_comments=False,
        include_tables=False,
        favor_precision=True,
        deduplicate=True,
    )
    text = normalize_whitespace(extracted or "")
    if len(text) < 280:
        return text, "extract_short"
    return text, "ok"


def main() -> None:
    wb = load_workbook(INPUT_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    raw_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = (row[0] or "").strip()
        link = (row[1] or "").strip()
        if title or link:
            raw_rows.append({"row": row_idx, "title": title, "link": link})

    stats = Counter()
    rows = []
    seen_titles = set()
    for item in raw_rows:
        title_key = normalize_key(item["title"])
        if title_key and title_key in seen_titles:
            stats['drop_duplicate_title'] += 1
            continue
        if title_key:
            seen_titles.add(title_key)
        rows.append(item)

    fetched = []
    for idx, item in enumerate(rows, start=1):
        content, status = fetch_article(item["link"])
        item["content"] = content
        item["fetch_status"] = status
        fetched.append(item)
        if idx % 20 == 0:
            print(f"fetched={idx}/{len(rows)}")

    seen_contents = set()
    kept = []
    for item in fetched:
        body_key = content_key(item["content"])
        if body_key and body_key in seen_contents:
            stats['drop_duplicate_content'] += 1
            continue
        if body_key:
            seen_contents.add(body_key)

        item['label'] = classify(item['title'], item['content'])
        kept.append(item)
        stats['kept'] += 1
        stats[f"label_{item['label']}"] += 1
        stats[f"fetch_{item['fetch_status']}"] += 1

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = SHEET_NAME
    out_ws.append(["Judul", "Link", "ISI BERITA", "Label", "STATUS EKSTRAKSI"])
    for item in kept:
        out_ws.append([item['title'], item['link'], item['content'], item['label'], item['fetch_status']])
    out_wb.save(OUTPUT_PATH)

    print(f"saved={OUTPUT_PATH}")
    print(f"original_rows={len(raw_rows)}")
    print(f"after_title_dedupe={len(rows)}")
    for key in ['drop_duplicate_title', 'drop_duplicate_content', 'kept', 'label_positif', 'label_netral', 'label_negatif']:
        print(f"{key}={stats[key]}")


if __name__ == "__main__":
    main()
