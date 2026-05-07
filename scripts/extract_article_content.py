from pathlib import Path
from time import sleep

import requests
import trafilatura
from openpyxl import load_workbook


INPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita.xlsx")
OUTPUT_PATH = Path(r"D:\PROGRAMMING\Analisis Sentimen\data\data-berita-dengan-isi.xlsx")
SHEET_NAME = "Data"
TITLE_HEADER = "Judul"
LINK_HEADER = "Link"
CONTENT_HEADER = "ISI BERITA"
STATUS_HEADER = "STATUS EKSTRAKSI"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/132.0.0.0 Safari/537.36"
)


def normalize_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    cleaned = [line for line in lines if line]
    return "\n\n".join(cleaned).strip()


def fetch_article(url: str) -> tuple[str, str]:
    try:
        response = requests.get(
            url,
            timeout=30,
            headers={"User-Agent": USER_AGENT},
        )
        response.raise_for_status()
    except Exception as exc:
        return "", f"request_error: {type(exc).__name__}"

    extracted = trafilatura.extract(
        response.text,
        url=url,
        include_comments=False,
        include_tables=False,
        favor_precision=True,
        deduplicate=True,
    )
    if not extracted:
        downloaded = trafilatura.fetch_url(url)
        if downloaded:
            extracted = trafilatura.extract(
                downloaded,
                url=url,
                include_comments=False,
                include_tables=False,
                favor_precision=True,
                deduplicate=True,
            )

    if not extracted:
        return "", "extract_error: empty_content"

    return normalize_text(extracted), "ok"


def main() -> None:
    workbook = load_workbook(INPUT_PATH)
    sheet = workbook[SHEET_NAME]

    headers = [sheet.cell(row=1, column=idx).value for idx in range(1, sheet.max_column + 1)]
    if TITLE_HEADER not in headers or LINK_HEADER not in headers:
        raise ValueError("Header sheet 'Data' tidak sesuai ekspektasi.")

    title_col = headers.index(TITLE_HEADER) + 1
    link_col = headers.index(LINK_HEADER) + 1
    content_col = sheet.max_column + 1
    status_col = sheet.max_column + 2

    sheet.cell(row=1, column=content_col).value = CONTENT_HEADER
    sheet.cell(row=1, column=status_col).value = STATUS_HEADER

    total = 0
    success = 0
    failed = 0

    for row_idx in range(2, sheet.max_row + 1):
        title = sheet.cell(row=row_idx, column=title_col).value
        url = sheet.cell(row=row_idx, column=link_col).value
        if not title and not url:
            continue

        total += 1
        content, status = fetch_article(str(url).strip())
        sheet.cell(row=row_idx, column=content_col).value = content
        sheet.cell(row=row_idx, column=status_col).value = status

        if status == "ok" and content:
            success += 1
        else:
            failed += 1

        if row_idx % 20 == 0:
            print(f"processed={total} success={success} failed={failed}")
        sleep(0.3)

    workbook.save(OUTPUT_PATH)
    print(f"saved={OUTPUT_PATH}")
    print(f"total={total}")
    print(f"success={success}")
    print(f"failed={failed}")


if __name__ == "__main__":
    main()
