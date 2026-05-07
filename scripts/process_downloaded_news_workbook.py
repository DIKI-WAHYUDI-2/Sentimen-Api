from __future__ import annotations

import argparse
import re
import shutil
import threading
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Iterable
from urllib.parse import urlsplit, urlunsplit

import requests
import trafilatura
from bs4 import BeautifulSoup
from newspaper import Article
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/135.0.0.0 Safari/537.36"
)
REQUEST_HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
}

VALIDASI_JUDUL = "Validasi Judul"
VALIDASI_LINK = "Validasi Link"
STATUS_DUPLIKASI = "Status Duplikasi"
ISI_BERITA = "Isi Berita Hasil Scrape"
STATUS_SCRAPE = "Status Scrape"

thread_local = threading.local()


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    normalized = " ".join(str(value).strip().lower().split())
    return normalized or None


def normalize_url(value: object) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    parts = urlsplit(raw)
    path = parts.path.rstrip("/") or "/"
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), path, parts.query, ""))


def clean_text(text: str | None) -> str:
    if not text:
        return ""
    text = text.replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" ?\n ?", "\n", text)
    return text.strip()


def get_session() -> requests.Session:
    session = getattr(thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(REQUEST_HEADERS)
        thread_local.session = session
    return session


def extract_with_bs4(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "iframe", "svg", "form", "button", "header", "footer"]):
        tag.decompose()

    selectors = [
        "article",
        "[itemprop='articleBody']",
        ".article-content",
        ".post-content",
        ".entry-content",
        ".td-post-content",
        ".single-content",
        ".read__content",
        ".content-detail",
        ".detail-content",
        ".news-content",
        ".post-body",
        ".content",
        ".artikel-content",
        ".main-content",
    ]

    candidates: list[str] = []
    for selector in selectors:
        for node in soup.select(selector):
            text = clean_text(node.get_text("\n"))
            if len(text) > 400:
                candidates.append(text)

    if candidates:
        return max(candidates, key=len)

    paragraphs = []
    for paragraph in soup.find_all("p"):
        text = clean_text(paragraph.get_text(" "))
        if len(text) > 40:
            paragraphs.append(text)
    return clean_text("\n\n".join(paragraphs))


def fetch_article(url: str) -> tuple[str, str]:
    try:
        response = get_session().get(url, timeout=25, allow_redirects=True)
        response.raise_for_status()
    except Exception as exc:
        return "", f"request_error:{type(exc).__name__}"

    html = response.text
    text = ""

    try:
        text = clean_text(
            trafilatura.extract(
                html,
                url=url,
                include_comments=False,
                include_tables=False,
                favor_precision=True,
                deduplicate=True,
            )
        )
    except Exception:
        text = ""

    if len(text) < 400:
        try:
            article = Article(url="")
            article.set_html(html)
            article.parse()
            article_text = clean_text(article.text)
            if len(article_text) > len(text):
                text = article_text
        except Exception:
            pass

    if len(text) < 400:
        try:
            soup_text = extract_with_bs4(html)
            if len(soup_text) > len(text):
                text = soup_text
        except Exception:
            pass

    if not text:
        return "", "extract_error:empty_content"
    if len(text) < 280:
        return text, "extract_short"
    return text, "ok"


def ensure_output_file(input_path: Path, output_path: Path) -> None:
    if not output_path.exists():
        shutil.copy2(input_path, output_path)


def ensure_headers(sheet) -> dict[str, int]:
    headers = {sheet.cell(1, col).value: col for col in range(1, sheet.max_column + 1)}
    for header in [VALIDASI_JUDUL, VALIDASI_LINK, STATUS_DUPLIKASI, ISI_BERITA, STATUS_SCRAPE]:
        if header not in headers:
            new_col = sheet.max_column + 1
            sheet.cell(1, new_col).value = header
            headers[header] = new_col

    widths = {
        headers[VALIDASI_JUDUL]: 18,
        headers[VALIDASI_LINK]: 18,
        headers[STATUS_DUPLIKASI]: 22,
        headers[ISI_BERITA]: 110,
        headers[STATUS_SCRAPE]: 32,
    }
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    return headers


def build_row_data(sheet, headers: dict[str, int]) -> tuple[list[dict], Counter, Counter]:
    rows: list[dict] = []
    title_counter: Counter = Counter()
    link_counter: Counter = Counter()

    for row_idx in range(2, sheet.max_row + 1):
        title_raw = sheet.cell(row_idx, headers["Judul"]).value
        link_raw = sheet.cell(row_idx, headers["Link"]).value
        title = normalize_text(title_raw)
        link = normalize_url(link_raw)
        rows.append({"row": row_idx, "title": title, "link": link})
        if title:
            title_counter[title] += 1
        if link:
            link_counter[link] += 1
    return rows, title_counter, link_counter


def write_validation(sheet, rows: Iterable[dict], headers: dict[str, int], title_counter: Counter, link_counter: Counter) -> None:
    for item in rows:
        title_dup = bool(item["title"] and title_counter[item["title"]] > 1)
        link_dup = bool(item["link"] and link_counter[item["link"]] > 1)

        sheet.cell(item["row"], headers[VALIDASI_JUDUL]).value = "Kosong" if not item["title"] else ("Duplikat" if title_dup else "Unik")
        sheet.cell(item["row"], headers[VALIDASI_LINK]).value = "Kosong" if not item["link"] else ("Duplikat" if link_dup else "Unik")

        if title_dup and link_dup:
            status = "Duplikat Judul+Link"
        elif title_dup:
            status = "Duplikat Judul"
        elif link_dup:
            status = "Duplikat Link"
        elif not item["link"]:
            status = "Link Kosong"
        else:
            status = "Unik"
        sheet.cell(item["row"], headers[STATUS_DUPLIKASI]).value = status


def fill_existing_duplicate_rows(sheet, rows_by_link: dict[str, list[int]], headers: dict[str, int]) -> int:
    copied = 0
    for row_indexes in rows_by_link.values():
        cached_content = ""
        cached_status = ""
        for row_idx in row_indexes:
            status = sheet.cell(row_idx, headers[STATUS_SCRAPE]).value
            content = sheet.cell(row_idx, headers[ISI_BERITA]).value
            if status and str(status).strip():
                cached_status = str(status)
                cached_content = str(content or "")
                break
        if not cached_status:
            continue
        for row_idx in row_indexes:
            if not sheet.cell(row_idx, headers[STATUS_SCRAPE]).value:
                sheet.cell(row_idx, headers[STATUS_SCRAPE]).value = cached_status
                sheet.cell(row_idx, headers[ISI_BERITA]).value = cached_content
                copied += 1
    return copied


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--sheet", default="Data")
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--save-every", type=int, default=50)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    ensure_output_file(input_path, output_path)
    workbook = load_workbook(output_path)
    sheet = workbook[args.sheet]
    headers = ensure_headers(sheet)

    if "Judul" not in headers or "Link" not in headers:
        raise ValueError("Sheet target harus memiliki header 'Judul' dan 'Link'.")

    rows, title_counter, link_counter = build_row_data(sheet, headers)
    write_validation(sheet, rows, headers, title_counter, link_counter)

    rows_by_link: dict[str, list[int]] = defaultdict(list)
    for item in rows:
        if item["link"]:
            rows_by_link[item["link"]].append(item["row"])
        else:
            sheet.cell(item["row"], headers[STATUS_SCRAPE]).value = "Link kosong"

    copied = fill_existing_duplicate_rows(sheet, rows_by_link, headers)
    workbook.save(output_path)

    pending_links: list[str] = []
    for link, row_indexes in rows_by_link.items():
        if any(not sheet.cell(row_idx, headers[STATUS_SCRAPE]).value for row_idx in row_indexes):
            pending_links.append(link)

    print(f"rows_total={len(rows)}")
    print(f"unique_links={len(rows_by_link)}")
    print(f"duplicate_title_rows={sum(1 for item in rows if item['title'] and title_counter[item['title']] > 1)}")
    print(f"duplicate_link_rows={sum(1 for item in rows if item['link'] and link_counter[item['link']] > 1)}")
    print(f"copied_from_existing={copied}")
    print(f"pending_links={len(pending_links)}")

    completed = 0
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        future_to_link = {executor.submit(fetch_article, link): link for link in pending_links}
        for future in as_completed(future_to_link):
            link = future_to_link[future]
            try:
                content, status = future.result()
            except Exception as exc:
                content, status = "", f"worker_error:{type(exc).__name__}"

            for row_idx in rows_by_link[link]:
                sheet.cell(row_idx, headers[ISI_BERITA]).value = content
                sheet.cell(row_idx, headers[STATUS_SCRAPE]).value = status

            completed += 1
            if completed % args.save_every == 0:
                workbook.save(output_path)
                print(f"saved_progress={completed}/{len(pending_links)}")

    workbook.save(output_path)

    ok_rows = 0
    failed_rows = 0
    for item in rows:
        status = sheet.cell(item["row"], headers[STATUS_SCRAPE]).value
        if status == "ok":
            ok_rows += 1
        elif status:
            failed_rows += 1

    print(f"saved={output_path}")
    print(f"ok_rows={ok_rows}")
    print(f"non_ok_rows={failed_rows}")


if __name__ == "__main__":
    main()
