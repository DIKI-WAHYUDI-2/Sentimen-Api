from collections import defaultdict
from datetime import date, datetime
import logging
from pathlib import Path

import pandas as pd

from apps.models.news import News
from apps.repositories.news_repository import NewsRepository
from apps.utils.indobert_inference import analyze_sentiment
from apps.utils.scraper import get_news

import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import io

logger = logging.getLogger(__name__)


class NewsService:
    BASE_DIR = Path(__file__).resolve().parents[2]
    EXPORT_DIR = BASE_DIR / "data" / "exports"
    MONTH_NAMES = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December",
    }

    @staticmethod
    def _parse_date(value):
        if value is None or value == "":
            return None

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except ValueError as exc:
                raise ValueError("Date format must be YYYY-MM-DD") from exc

        raise ValueError("Invalid date value")

    @staticmethod
    def normalize_news_payload(data):
        return {
            "title": data.get("title"),
            "published_at": data.get("published_at"),
            "source": data.get("source"),
            "url": data.get("url"),
            "content": data.get("content"),
            "sentiment": data.get("sentiment"),
        }

    @staticmethod
    def get_all_news():
        return NewsRepository.find_all()

    @staticmethod
    def get_paginated_news(page, limit, sentiment=None):
        return NewsRepository.find_paginated(page, limit, sentiment=sentiment)

    @staticmethod
    def create_news(data):
        payload = NewsService.normalize_news_payload(data)
        required_fields = ("title", "published_at", "source", "url", "sentiment")
        missing_fields = [field for field in required_fields if not payload.get(field)]
        if missing_fields:
            missing_field_list = ", ".join(missing_fields)
            raise ValueError(f"Missing required fields")

        news_item = News(
            title=payload.get("title"),
            published_at=NewsService._parse_date(payload.get("published_at")),
            source=payload.get("source"),
            url=payload.get("url"),
            content=payload.get("content"),
            sentiment=payload.get("sentiment"),
        )
        return NewsRepository.save(news_item)

    @staticmethod
    def update_news(news_id, data):
        news_item = NewsRepository.find_by_id(news_id)
        if not news_item:
            return None

        payload = NewsService.normalize_news_payload(data)
        news_item.title = payload.get("title", news_item.title)
        if payload.get("published_at") is not None:
            news_item.published_at = NewsService._parse_date(payload.get("published_at"))
        news_item.source = payload.get("source", news_item.source)
        news_item.url = payload.get("url", news_item.url)
        news_item.content = payload.get("content", news_item.content)
        news_item.sentiment = payload.get("sentiment", news_item.sentiment)
        return NewsRepository.save(news_item)

    @staticmethod
    def delete_news(news_id):
        news_item = NewsRepository.find_by_id(news_id)
        if not news_item:
            return False
        return NewsRepository.delete(news_id) is not None

    @staticmethod
    def _export_scrape_results_to_excel(items, start_date=None, end_date=None):
        if not items:
            return None

        NewsService.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        if start_date and end_date:
            filename = f"scraping_news_{start_date}_{end_date}.xlsx"
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"scraping_news_without_date_filter_{timestamp}.xlsx"
        export_path = NewsService.EXPORT_DIR / filename

        dataframe = pd.DataFrame(items)
        ordered_columns = [
            "title",
            "date",
            "source",
            "url",
            "content",
            "sentiment",
        ]
        available_columns = [column for column in ordered_columns if column in dataframe.columns]
        if available_columns:
            dataframe = dataframe[available_columns]

        dataframe.to_excel(export_path, index=False)
        logger.info("Scraped news exported to Excel", extra={"path": str(export_path)})
        return str(export_path)

    @staticmethod
    def export_scraped_news(items, start_date=None, end_date=None):
        normalized_items = []
        for item in items:
            normalized_items.append(
                {
                    "title": item.get("title"),
                    "date": item.get("date", item.get("published_at")),
                    "source": item.get("source"),
                    "url": item.get("url"),
                    "content": item.get("content"),
                    "sentiment": item.get("sentiment"),
                }
            )
        return NewsService._export_scrape_results_to_excel(
            normalized_items,
            start_date=start_date,
            end_date=end_date,
        )

    @staticmethod
    def scrape_classify_and_save(data):
        news_items = get_news(data)
        if not news_items:
            logger.info("Scrape returned no items")
            return {"count": 0, "saved_count": 0, "skipped_count": 0, "failed_count": 0, "data": []}

        saved_items = []
        skipped_count = 0
        failed_count = 0

        for scraped_item in news_items:
            try:
                url = scraped_item.get("url", "#")

                if NewsRepository.find_by_url(url):
                    skipped_count += 1
                    logger.debug("Skipping duplicate news item", extra={"url": url})
                    continue

                published_at = scraped_item.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
                content = scraped_item.get("content") or ""
                title = scraped_item.get("title", "Untitled")

                archived_item = {
                    "title": title,
                    "date": published_at,
                    "source": scraped_item.get("source", "Unknown source"),
                    "url": url,
                    "content": content,
                    "sentiment": analyze_sentiment(title, content),
                }

                payload = {
                    "title": archived_item["title"],
                    "published_at": archived_item["date"],
                    "source": archived_item["source"],
                    "url": archived_item["url"],
                    "content": archived_item["content"],
                    "sentiment": archived_item["sentiment"],
                }
                if NewsRepository.create(payload):
                    saved_items.append(archived_item)
                else:
                    failed_count += 1
            except Exception:
                failed_count += 1
                logger.exception("Failed to process scraped news item")

        logger.info(
            "Scrape pipeline completed",
            extra={"saved": len(saved_items), "skipped": skipped_count, "failed": failed_count},
        )
        export_file = NewsService._export_scrape_results_to_excel(saved_items)
        return {
            "count": len(news_items),
            "saved_count": len(saved_items),
            "skipped_count": skipped_count,
            "failed_count": failed_count,
            "data": saved_items,
            "export_file": export_file,
        }

    MONTH_NAMES_ID = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
        5: "Mei", 6: "Jun", 7: "Jul", 8: "Agt",
        9: "Sep", 10: "Okt", 11: "Nov", 12: "Des",
    }

    @staticmethod
    def export_news_by_month_range(start_month_str, end_month_str=None):
        

        try:
            start_dt = datetime.strptime(start_month_str, "%Y-%m")
        except ValueError as exc:
            raise ValueError("Format bulan mulai harus YYYY-MM") from exc

        if end_month_str:
            try:
                end_dt = datetime.strptime(end_month_str, "%Y-%m")
            except ValueError as exc:
                raise ValueError("Format bulan akhir harus YYYY-MM") from exc
        else:
            end_dt = start_dt

        if end_dt < start_dt:
            raise ValueError("Bulan akhir tidak boleh lebih kecil dari bulan mulai")

        months = []
        cursor = start_dt
        while cursor <= end_dt:
            months.append((cursor.year, cursor.month))
            if cursor.month == 12:
                cursor = cursor.replace(year=cursor.year + 1, month=1)
            else:
                cursor = cursor.replace(month=cursor.month + 1)

        if len(months) > 12:
            raise ValueError("Rentang maksimal adalah 12 bulan")

        news_items = NewsRepository.find_by_month_range(
            start_dt.year, start_dt.month, end_dt.year, end_dt.month
        )

        grouped = {(y, m): [] for y, m in months}
        for item in news_items:
            key = (item.published_at.year, item.published_at.month)
            if key in grouped:
                grouped[key].append(item)

        template_path = NewsService.BASE_DIR / "data" / "Template-Export.xlsx"
        if not template_path.exists():
            raise FileNotFoundError(f"Template tidak ditemukan: {template_path}")

        template_wb = load_workbook(template_path)
        template_ws = template_wb.active

        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

        month_names_id = {
            1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
            5: "Mei", 6: "Jun", 7: "Jul", 8: "Agt",
            9: "Sep", 10: "Okt", 11: "Nov", 12: "Des",
        }

        for year, month in months:
            sheet_name = f"{month_names_id[month]} {year}"
            ws = wb.create_sheet(title=sheet_name)

            for row in template_ws.iter_rows():
                for cell in row:
                    new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            for col_letter, col_dim in template_ws.column_dimensions.items():
                ws.column_dimensions[col_letter].width = col_dim.width

            for row_num, row_dim in template_ws.row_dimensions.items():
                ws.row_dimensions[row_num].height = row_dim.height

            for merged_range in template_ws.merged_cells.ranges:
                ws.merge_cells(str(merged_range))

            data_start_row = 3
            template_data_end_row = template_ws.max_row

            def copy_template_row_style(source_row, target_row):
                ws.row_dimensions[target_row].height = template_ws.row_dimensions[source_row].height
                for column in range(1, template_ws.max_column + 1):
                    source_cell = template_ws.cell(row=source_row, column=column)
                    target_cell = ws.cell(row=target_row, column=column)
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)

            items = grouped[(year, month)]
            if not items:
                ws.cell(row=data_start_row, column=1, value="Tidak ada data")
            else:
                for idx, item in enumerate(items, start=1):
                    row_num = data_start_row + idx - 1
                    if row_num > template_data_end_row:
                        copy_template_row_style(template_data_end_row, row_num)
                    published_str = item.published_at.strftime("%d-%m-%Y") if item.published_at else ""
                    ws.cell(row=row_num, column=1, value=idx)
                    ws.cell(row=row_num, column=2, value=published_str)
                    ws.cell(row=row_num, column=3, value=item.title or "")
                    ws.cell(row=row_num, column=4, value=item.url or "")
                    ws.cell(row=row_num, column=5, value=item.source or "")
                    ws.cell(row=row_num, column=6, value=item.sentiment or "")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def search_news(keyword):
        return NewsRepository.find_by_keyword(keyword)

    @staticmethod
    def get_sentiment_summary():
        return NewsRepository.get_sentiment_data()

    @staticmethod
    def get_monthly_sentiment_summary(year=None):
        available_years = NewsRepository.get_available_years()
        if not available_years:
            return {"year": None, "years": [], "data": []}

        selected_year = year if year in available_years else available_years[0]
        rows = NewsRepository.get_sentiment_and_date_by_year(selected_year)
        monthly_sentiment = {
            month: {"positive": 0, "negative": 0, "neutral": 0}
            for month in range(1, 13)
        }

        for published_at, sentiment in rows:
            month = published_at.month
            normalized_sentiment = sentiment.strip().lower()
            if normalized_sentiment == "positif":
                normalized_sentiment = "positive"
            elif normalized_sentiment == "negatif":
                normalized_sentiment = "negative"
            elif normalized_sentiment == "netral":
                normalized_sentiment = "neutral"

            if normalized_sentiment not in monthly_sentiment[month]:
                normalized_sentiment = "neutral"

            monthly_sentiment[month][normalized_sentiment] += 1

        return {
            "year": selected_year,
            "years": available_years,
            "data": [
                {
                    "month": NewsService.MONTH_NAMES[month],
                    "positive": counts["positive"],
                    "neutral": counts["neutral"],
                    "negative": counts["negative"],
                    "bulan": NewsService.MONTH_NAMES[month],
                    "positif": counts["positive"],
                    "netral": counts["neutral"],
                    "negatif": counts["negative"],
                }
                for month, counts in monthly_sentiment.items()
            ],
        }
